"""
Gestão de Dívidas de Fornecedores
----------------------------------
App Streamlit para acompanhar o status das dívidas com clientes/fornecedores:
quanto se deve, quanto já foi pago, quanto resta, e quantas dívidas já foram
quitadas versus quantas ainda estão ativas.

Fonte de dados padrão: aba "RESUMO" da planilha de Contas a Receber.
"""

import os
import html
from io import BytesIO
from datetime import datetime, date
from utils.Functions import rodape
import pandas as pd
import streamlit as st

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)

# ---------------------------------------------------------------------------
# Configuração da página
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Gestão de Dívidas | Fornecedores",
    page_icon="💧",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# Paleta de cores — Verde neon água
# ---------------------------------------------------------------------------
PRIMARY       = "#00E6B8"   # verde neon água
PRIMARY_DARK  = "#00B89C"
PRIMARY_DEEP  = "#017F6B"   # usado em cabeçalhos (tabela / pdf / excel)
PRIMARY_SOFT  = "#EAFBF6"   # fundo suave
DANGER        = "#FF6B6B"   # dívidas ativas
DANGER_DARK   = "#E64A4A"
TEXT_DARK     = "#0B2B26"
TEXT_MUTED    = "#5C7A75"
BORDER_SOFT   = "#D2F3EA"

DATA_PATH = os.path.join(os.path.dirname(__file__), "data", "Contas_a_receber.xlsx")

# ---------------------------------------------------------------------------
# CSS customizado
# ---------------------------------------------------------------------------
CUSTOM_CSS = f"""
<style>
    .block-container {{ padding-top: 1.6rem; padding-bottom: 3rem; }}

    /* ---------- Cabeçalho ---------- */
    .app-header {{
        background: linear-gradient(120deg, {PRIMARY_DEEP} 0%, {PRIMARY} 100%);
        border-radius: 18px;
        padding: 28px 34px;
        margin-bottom: 26px;
        color: #ffffff;
        box-shadow: 0 10px 30px rgba(0, 168, 143, 0.28);
    }}
    .app-header h1 {{
        margin: 0;
        font-size: 28px;
        font-weight: 800;
        letter-spacing: -0.02em;
    }}
    .app-header p {{
        margin: 6px 0 0 0;
        font-size: 14.5px;
        opacity: 0.92;
    }}

    /* ---------- Cards KPI ---------- */
    .kpi-card {{
        position: relative;
        background: linear-gradient(155deg, #ffffff 0%, {PRIMARY_SOFT} 120%);
        border: 1px solid {BORDER_SOFT};
        border-radius: 16px;
        padding: 18px 20px 16px 24px;
        box-shadow: 0 4px 16px rgba(0, 184, 148, 0.12);
        overflow: hidden;
        height: 118px;
    }}
    .kpi-card::before {{
        content: "";
        position: absolute;
        top: 0; left: 0; bottom: 0;
        width: 6px;
        background: linear-gradient(180deg, {PRIMARY}, {PRIMARY_DARK});
    }}
    .kpi-card.danger::before {{
        background: linear-gradient(180deg, {DANGER}, {DANGER_DARK});
    }}
    .kpi-top {{ display: flex; align-items: center; justify-content: space-between; }}
    .kpi-icon {{ font-size: 22px; opacity: 0.9; }}
    .kpi-label {{
        font-size: 12.5px;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.04em;
        color: {TEXT_MUTED};
        margin-top: 6px;
    }}
    .kpi-value {{
        font-size: 25px;
        font-weight: 800;
        color: {TEXT_DARK};
        margin-top: 2px;
        white-space: nowrap;
    }}
    .kpi-card.danger .kpi-value {{ color: {DANGER_DARK}; }}
    .kpi-sub {{ font-size: 11.5px; color: {TEXT_MUTED}; margin-top: 2px; }}

    /* ---------- Badges de status ---------- */
    .badge {{
        display: inline-block;
        padding: 4px 13px;
        border-radius: 20px;
        font-size: 12px;
        font-weight: 700;
        letter-spacing: 0.02em;
        white-space: nowrap;
    }}
    .badge-quitada {{
        background: rgba(0, 230, 184, 0.14);
        color: {PRIMARY_DEEP};
        border: 1px solid {PRIMARY};
    }}
    .badge-ativa {{
        background: rgba(255, 107, 107, 0.12);
        color: {DANGER_DARK};
        border: 1px solid {DANGER};
    }}

    /* ---------- Barra de progresso de pagamento ---------- */
    .progress-wrap {{ display: flex; align-items: center; gap: 8px; }}
    .progress-track {{
        background: #E7F8F2;
        border-radius: 8px;
        height: 9px;
        width: 110px;
        overflow: hidden;
        display: inline-block;
    }}
    .progress-fill {{
        height: 100%;
        border-radius: 8px;
        background: linear-gradient(90deg, {PRIMARY_DARK}, {PRIMARY});
    }}
    .progress-pct {{ font-size: 12px; color: {TEXT_MUTED}; font-weight: 600; min-width: 38px; }}

    /* ---------- Tabela customizada ---------- */
    .table-wrapper {{
        border-radius: 16px;
        overflow: hidden;
        border: 1px solid {BORDER_SOFT};
        box-shadow: 0 6px 22px rgba(0, 168, 143, 0.14);
        max-height: 600px;
        overflow-y: auto;
        margin-top: 4px;
    }}
    table.data-table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 13.5px;
        font-family: inherit;
    }}
    table.data-table thead th {{
        background: linear-gradient(120deg, {PRIMARY_DEEP}, {PRIMARY});
        color: #ffffff;
        text-align: left;
        padding: 13px 16px;
        font-weight: 700;
        font-size: 12.5px;
        text-transform: uppercase;
        letter-spacing: 0.03em;
        position: sticky;
        top: 0;
        z-index: 2;
    }}
    table.data-table tbody td {{
        padding: 11px 16px;
        border-bottom: 1px solid #EAF7F2;
        color: {TEXT_DARK};
    }}
    table.data-table tbody tr:nth-child(even) {{ background: #F6FFFC; }}
    table.data-table tbody tr:hover {{ background: #E8FFF7; }}
    .cell-money {{ font-weight: 600; white-space: nowrap; }}
    .cell-muted {{ color: {TEXT_MUTED}; }}

    /* ---------- Botões de download ---------- */
    div[data-testid="stDownloadButton"] button {{
        background: linear-gradient(120deg, {PRIMARY}, {PRIMARY_DARK});
        color: #ffffff;
        border: none;
        border-radius: 10px;
        padding: 10px 22px;
        font-weight: 700;
        box-shadow: 0 4px 14px rgba(0, 184, 148, 0.32);
        transition: transform 0.12s ease, box-shadow 0.12s ease;
    }}
    div[data-testid="stDownloadButton"] button:hover {{
        transform: translateY(-1px);
        box-shadow: 0 6px 18px rgba(0, 184, 148, 0.45);
        color: #ffffff;
    }}
    .export-card {{
        background: {PRIMARY_SOFT};
        border: 1px solid {BORDER_SOFT};
        border-radius: 16px;
        padding: 20px 24px;
        margin-top: 10px;
    }}
    .export-card h4 {{ margin: 0 0 4px 0; color: {TEXT_DARK}; }}
    .export-card p {{ margin: 0 0 14px 0; color: {TEXT_MUTED}; font-size: 13px; }}

    .section-title {{
        font-size: 18px;
        font-weight: 800;
        color: {TEXT_DARK};
        margin: 6px 0 12px 0;
    }}
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Funções utilitárias de formatação
# ---------------------------------------------------------------------------
def fmt_brl(valor: float) -> str:
    """Formata número como moeda brasileira: R$ 1.234,56"""
    if pd.isna(valor):
        valor = 0.0
    texto = f"{valor:,.2f}"
    texto = texto.replace(",", "§").replace(".", ",").replace("§", ".")
    return f"R$ {texto}"


def fmt_data(valor) -> str:
    """Formata datas para dd/mm/aaaa; mantém texto (ex: 'DESCREDÊNCIADA') como está."""
    if isinstance(valor, (pd.Timestamp, datetime, date)):
        return valor.strftime("%d/%m/%Y")
    if pd.isna(valor):
        return "-"
    return str(valor)


def pct_pago(divida: float, pago: float) -> float:
    if not divida or divida <= 0:
        return 100.0 if pago > 0 else 0.0
    return max(0.0, min(100.0, (pago / divida) * 100))


# ---------------------------------------------------------------------------
# Carregamento de dados
# ---------------------------------------------------------------------------
REQUIRED_COLS = [
    "Nome do cliente", "VALOR DIVIDA", "VALOR PAGO",
    "VALOR RESTANTE", "TERMINO DA DIVIDA", "QUITADO",
]


@st.cache_data(show_spinner=False)
def carregar_dados(file_bytes_or_path) -> pd.DataFrame:
    df = pd.read_excel(file_bytes_or_path, sheet_name="RESUMO")
    df.columns = [c.strip() for c in df.columns]
    faltantes = [c for c in REQUIRED_COLS if c not in df.columns]
    if faltantes:
        raise ValueError(
            "A aba 'RESUMO' não contém as colunas esperadas: " + ", ".join(faltantes)
        )
    df["QUITADO"] = df["QUITADO"].astype(str).str.strip().str.upper()
    for col in ["VALOR DIVIDA", "VALOR PAGO", "VALOR RESTANTE"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    return df


# ---------------------------------------------------------------------------
# Exportação para Excel (.xlsx) — mantendo o padrão visual do app
# ---------------------------------------------------------------------------
def gerar_excel(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    df_export = df.copy()
    df_export["TERMINO DA DIVIDA"] = df_export["TERMINO DA DIVIDA"].apply(fmt_data)
    df_export["STATUS"] = df_export["QUITADO"].apply(lambda v: "QUITADA" if v == "SIM" else "ATIVA")
    df_export = df_export.drop(columns=["QUITADO"])
    df_export = df_export.rename(columns={
        "Nome do cliente": "Cliente",
        "VALOR DIVIDA": "Valor da Dívida",
        "VALOR PAGO": "Valor Pago",
        "VALOR RESTANTE": "Valor Restante",
        "TERMINO DA DIVIDA": "Término da Dívida",
    })

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Dívidas Fornecedores")
        ws = writer.sheets["Dívidas Fornecedores"]

        header_fill = PatternFill("solid", fgColor="017F6B")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        col_idx = {c: i + 1 for i, c in enumerate(df_export.columns)}
        money_cols = ["Valor da Dívida", "Valor Pago", "Valor Restante"]
        green_fill = PatternFill("solid", fgColor="D2F3EA")
        red_fill = PatternFill("solid", fgColor="FFE1E1")

        for row in range(2, ws.max_row + 1):
            for c in money_cols:
                cell = ws.cell(row=row, column=col_idx[c])
                cell.number_format = '"R$" #,##0.00'
                cell.alignment = Alignment(horizontal="right")
            status_cell = ws.cell(row=row, column=col_idx["STATUS"])
            status_cell.alignment = center
            status_cell.font = Font(bold=True, color="017F6B" if status_cell.value == "QUITADA" else "E64A4A")
            status_cell.fill = green_fill if status_cell.value == "QUITADA" else red_fill
            ws.cell(row=row, column=col_idx["Cliente"]).alignment = left
            ws.cell(row=row, column=col_idx["Término da Dívida"]).alignment = center

        widths = {
            "Cliente": 42, "Valor da Dívida": 18, "Valor Pago": 16,
            "Valor Restante": 16, "Término da Dívida": 16, "STATUS": 12,
        }
        for col_name, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx[col_name])].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 22

        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = False

    return output.getvalue()


# ---------------------------------------------------------------------------
# Exportação para PDF — mantendo o padrão visual do app
# ---------------------------------------------------------------------------
def gerar_pdf(df: pd.DataFrame, kpis: dict) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        topMargin=1.4 * cm, bottomMargin=1.4 * cm,
        leftMargin=1.4 * cm, rightMargin=1.4 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TituloRelatorio", parent=styles["Heading1"],
        textColor=colors.HexColor(PRIMARY_DEEP), fontSize=18, spaceAfter=2,
    )
    sub_style = ParagraphStyle(
        "Subtitulo", parent=styles["Normal"],
        textColor=colors.HexColor(TEXT_MUTED), fontSize=10,
    )
    resumo_style = ParagraphStyle(
        "Resumo", parent=styles["Normal"],
        textColor=colors.HexColor(TEXT_DARK), fontSize=10.5, leading=15,
    )

    elementos = [
        Paragraph("Relatório de Dívidas de Fornecedores", title_style),
        Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", sub_style),
        Spacer(1, 10),
        HRFlowable(width="100%", thickness=1.2, color=colors.HexColor(PRIMARY)),
        Spacer(1, 10),
    ]

    resumo_html = (
        f"<b>Valor total em dívida:</b> {fmt_brl(kpis['total_divida'])} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Total já pago:</b> {fmt_brl(kpis['total_pago'])} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Total restante:</b> {fmt_brl(kpis['total_restante'])}<br/>"
        f"<b>Dívidas quitadas:</b> {kpis['qtd_quitadas']} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Dívidas ativas:</b> {kpis['qtd_ativas']} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>% pago do total:</b> {kpis['pct_pago_geral']:.1f}%"
    )
    elementos.append(Paragraph(resumo_html, resumo_style))
    elementos.append(Spacer(1, 16))

    cell_style = ParagraphStyle(
        "Celula", fontName="Helvetica", fontSize=8.5, leading=10.5,
        textColor=colors.HexColor(TEXT_DARK),
    )

    header = ["Cliente", "Valor Dívida", "Valor Pago", "Valor Restante", "Término", "Status"]
    dados = [header]
    status_rows = []  # guarda (linha_idx, status) para colorir depois
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        status = "Quitada" if row["QUITADO"] == "SIM" else "Ativa"
        status_rows.append((i, status))
        dados.append([
            Paragraph(html.escape(str(row["Nome do cliente"])), cell_style),
            fmt_brl(row["VALOR DIVIDA"]),
            fmt_brl(row["VALOR PAGO"]),
            fmt_brl(row["VALOR RESTANTE"]),
            fmt_data(row["TERMINO DA DIVIDA"]),
            status,
        ])

    tabela = Table(
        dados, repeatRows=1,
        colWidths=[8.2 * cm, 3.4 * cm, 3.4 * cm, 3.4 * cm, 3.0 * cm, 2.4 * cm],
    )

    estilo = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY_DEEP)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("ALIGN", (0, 1), (0, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(BORDER_SOFT)),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(PRIMARY_SOFT)]),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]
    for linha_idx, status in status_rows:
        cor = colors.HexColor(PRIMARY_DEEP) if status == "Quitada" else colors.HexColor(DANGER_DARK)
        estilo.append(("TEXTCOLOR", (5, linha_idx), (5, linha_idx), cor))
        estilo.append(("FONTNAME", (5, linha_idx), (5, linha_idx), "Helvetica-Bold"))

    tabela.setStyle(TableStyle(estilo))
    elementos.append(tabela)

    doc.build(elementos)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Cabeçalho
# ---------------------------------------------------------------------------
st.markdown(
    """
    <div class="app-header">
        <h1>💧 Gestão de Dívidas de Fornecedores</h1>
        <p>Acompanhamento de valores devidos, pagamentos realizados e status de quitação por fornecedor/cliente.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------------
# Sidebar — fonte de dados e filtros
# ---------------------------------------------------------------------------
with st.sidebar:
    st.markdown("## Base de dados")
    arquivo_upload = st.file_uploader(
        "Atualizar planilha (.xlsx com aba 'RESUMO')", type=["xlsx"]
    )
    st.caption("Se nenhum arquivo for enviado, a planilha padrão é utilizada.")

    st.divider()
    #st.markdown("---")
    st.markdown("## Filtros")

try:
    if arquivo_upload is not None:
        df_raw = carregar_dados(arquivo_upload)
    else:
        df_raw = carregar_dados(DATA_PATH)
except Exception as e:
    st.error(f"Não foi possível carregar os dados: {e}")
    st.stop()

with st.sidebar:
    status_opcao = st.radio(
        "Status da dívida", ["Todas", "Somente ativas", "Somente quitadas"], index=0
    )
    busca = st.text_input("Buscar por cliente", placeholder="Digite o nome do cliente...")
    st.markdown("---")
    st.caption(f"Total de registros na base: **{len(df_raw)}**")

df = df_raw.copy()
if status_opcao == "Somente ativas":
    df = df[df["QUITADO"] == "NÃO"]
elif status_opcao == "Somente quitadas":
    df = df[df["QUITADO"] == "SIM"]
if busca:
    df = df[df["Nome do cliente"].str.contains(busca, case=False, na=False)]

# ---------------------------------------------------------------------------
# KPIs
# ---------------------------------------------------------------------------
total_divida = df["VALOR DIVIDA"].sum()
total_pago = df["VALOR PAGO"].sum()
total_restante = df["VALOR RESTANTE"].sum()
qtd_quitadas = int((df["QUITADO"] == "SIM").sum())
qtd_ativas = int((df["QUITADO"] == "NÃO").sum())
qtd_total = len(df)
pct_pago_geral = pct_pago(total_divida, total_pago)

kpis = dict(
    total_divida=total_divida, total_pago=total_pago, total_restante=total_restante,
    qtd_quitadas=qtd_quitadas, qtd_ativas=qtd_ativas, pct_pago_geral=pct_pago_geral,
)

if df.empty:
    st.warning("Nenhum registro encontrado para os filtros selecionados.")
    st.stop()

st.divider()
st.markdown('<div class="section-title"> Visão geral</div>', unsafe_allow_html=True)

col1, col2, col3 = st.columns(3)
with col1:
    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-top"><span class="kpi-icon">💰</span></div>
            <div class="kpi-label">Valor total em dívida</div>
            <div class="kpi-value">{fmt_brl(total_divida)}</div>
        </div>
        """, unsafe_allow_html=True,
    )
with col2:
    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-top"><span class="kpi-icon">✅</span></div>
            <div class="kpi-label">Valor já pago</div>
            <div class="kpi-value">{fmt_brl(total_pago)}</div>
            <div class="kpi-sub">{pct_pago_geral:.1f}% do total da dívida</div>
        </div>
        """, unsafe_allow_html=True,
    )
with col3:
    st.markdown(
        f"""
        <div class="kpi-card danger">
            <div class="kpi-top"><span class="kpi-icon">⏳</span></div>
            <div class="kpi-label">Valor restante a receber</div>
            <div class="kpi-value">{fmt_brl(total_restante)}</div>
        </div>
        """, unsafe_allow_html=True,
    )

col4, col5, col6 = st.columns(3)
with col4:
    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-top"><span class="kpi-icon">🟢</span></div>
            <div class="kpi-label">Dívidas quitadas</div>
            <div class="kpi-value">{qtd_quitadas}</div>
            <div class="kpi-sub">de {qtd_total} registros ({(qtd_quitadas/qtd_total*100):.0f}%)</div>
        </div>
        """, unsafe_allow_html=True,
    )
with col5:
    st.markdown(
        f"""
        <div class="kpi-card danger">
            <div class="kpi-top"><span class="kpi-icon">🔴</span></div>
            <div class="kpi-label">Dívidas ativas</div>
            <div class="kpi-value">{qtd_ativas}</div>
            <div class="kpi-sub">de {qtd_total} registros ({(qtd_ativas/qtd_total*100):.0f}%)</div>
        </div>
        """, unsafe_allow_html=True,
    )
with col6:
    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-top"><span class="kpi-icon">📈</span></div>
            <div class="kpi-label">% pago do total</div>
            <div class="kpi-value">{pct_pago_geral:.1f}%</div>
            <div class="kpi-sub">considerando os registros filtrados</div>
        </div>
        """, unsafe_allow_html=True,
    )

st.divider()
st.write("")

# ---------------------------------------------------------------------------
# Tabela detalhada (HTML + CSS customizado)
# ---------------------------------------------------------------------------
st.markdown('<div class="section-title"> Detalhamento por cliente</div>', unsafe_allow_html=True)

linhas_html = []
df_ordenado = df.sort_values(by=["QUITADO", "VALOR RESTANTE"], ascending=[True, False])

for _, row in df_ordenado.iterrows():
    cliente = html.escape(str(row["Nome do cliente"]))
    quitado = row["QUITADO"] == "SIM"
    badge = (
        '<span class="badge badge-quitada">✓ Quitada</span>' if quitado
        else '<span class="badge badge-ativa">● Ativa</span>'
    )
    pct = pct_pago(row["VALOR DIVIDA"], row["VALOR PAGO"])
    progresso = (
        '<div class="progress-wrap">'
        f'<div class="progress-track"><div class="progress-fill" style="width:{pct:.0f}%;"></div></div>'
        f'<span class="progress-pct">{pct:.0f}%</span>'
        '</div>'
    )
    linhas_html.append(
        "<tr>"
        f"<td>{cliente}</td>"
        f'<td class="cell-money">{fmt_brl(row["VALOR DIVIDA"])}</td>'
        f'<td class="cell-money">{fmt_brl(row["VALOR PAGO"])}</td>'
        f'<td class="cell-money">{fmt_brl(row["VALOR RESTANTE"])}</td>'
        f"<td>{progresso}</td>"
        f'<td class="cell-muted">{fmt_data(row["TERMINO DA DIVIDA"])}</td>'
        f"<td>{badge}</td>"
        "</tr>"
    )

tabela_html = f"""
<div class="table-wrapper">
<table class="data-table">
    <thead>
        <tr>
            <th>Cliente</th>
            <th>Valor Dívida</th>
            <th>Valor Pago</th>
            <th>Valor Restante</th>
            <th>% Pago</th>
            <th>Término</th>
            <th>Status</th>
        </tr>
    </thead>
    <tbody>
        {''.join(linhas_html)}
    </tbody>
</table>
</div>
"""
st.markdown(tabela_html, unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Exportação de dados
# ---------------------------------------------------------------------------
st.divider()
st.markdown('<div class="section-title">Exportar dados</div>', unsafe_allow_html=True)

col_xlsx, col_pdf = st.columns(2)
data_hoje = datetime.now().strftime("%Y%m%d_%H%M")

with col_xlsx:
    excel_bytes = gerar_excel(df_ordenado)
    st.download_button(
        label="📊  Exportar para Excel",
        data=excel_bytes,
        file_name=f"dividas_fornecedores_{data_hoje}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

with col_pdf:
    pdf_bytes = gerar_pdf(df_ordenado, kpis)
    st.download_button(
        label="📄  Exportar para PDF",
        data=pdf_bytes,
        file_name=f"dividas_fornecedores_{data_hoje}.pdf",
        mime="application/pdf",
        use_container_width=True,
    )

rodape("RiaChuelo")
